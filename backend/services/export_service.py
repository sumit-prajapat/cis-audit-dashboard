"""
Export Service - Generate CSV, Excel, and PDF reports
"""
import csv
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT


class ExportService:
    """Service for exporting data in various formats"""
    
    @staticmethod
    def generate_csv(data: List[Dict[str, Any]], headers: List[str]) -> io.StringIO:
        """Generate CSV file from data"""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_excel(data: List[Dict[str, Any]], sheet_name: str = "Report") -> io.BytesIO:
        """Generate Excel file from data"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        if not data:
            return io.BytesIO()
        
        # Add headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                # Format datetime objects
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_scan_report_pdf(scan_data: Dict[str, Any], org_name: str) -> io.BytesIO:
        """Generate PDF report for scan results"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph(f"Security Compliance Report", title_style))
        story.append(Paragraph(f"<b>{org_name}</b>", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Report metadata
        metadata = [
            ['Report Date:', datetime.now().strftime("%B %d, %Y %H:%M")],
            ['Scan ID:', scan_data.get('id', 'N/A')],
            ['Device:', scan_data.get('device_hostname', 'N/A')],
            ['OS Type:', scan_data.get('device_os_type', 'N/A')],
        ]
        
        metadata_table = Table(metadata, colWidths=[2*inch, 4*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        story.append(metadata_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Summary section
        story.append(Paragraph("Executive Summary", heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Compliance Score', f"{scan_data.get('compliance_score', 0)}%"],
            ['Total Checks', str(scan_data.get('total_checks', 0))],
            ['Passed', str(scan_data.get('passed_checks', 0))],
            ['Failed', str(scan_data.get('failed_checks', 0))],
            ['Warnings', str(scan_data.get('warned_checks', 0))],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Findings severity
        story.append(Paragraph("Findings by Severity", heading_style))
        
        severity_data = [
            ['Severity', 'Count'],
            ['Critical', str(scan_data.get('critical_count', 0))],
            ['High', str(scan_data.get('high_count', 0))],
            ['Medium', str(scan_data.get('medium_count', 0))],
            ['Low', str(scan_data.get('low_count', 0))],
        ]
        
        severity_table = Table(severity_data, colWidths=[3*inch, 3*inch])
        severity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fee2e2')]),
        ]))
        
        story.append(severity_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Check details (if provided)
        if 'checks' in scan_data and scan_data['checks']:
            story.append(PageBreak())
            story.append(Paragraph("Detailed Check Results", heading_style))
            
            checks_data = [['Check ID', 'Title', 'Status', 'Severity']]
            
            for check in scan_data['checks'][:20]:  # Limit to first 20 checks
                checks_data.append([
                    check.get('check_id', '')[:15],
                    check.get('title', '')[:40],
                    check.get('status', ''),
                    check.get('severity', '')
                ])
            
            checks_table = Table(checks_data, colWidths=[1*inch, 3*inch, 1*inch, 1*inch])
            checks_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            
            story.append(checks_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_compliance_report_csv(scans: List[Dict[str, Any]]) -> io.StringIO:
        """Generate CSV report for compliance data"""
        headers = [
            'scan_id', 'device_hostname', 'device_os_type', 'scan_date',
            'compliance_score', 'total_checks', 'passed_checks', 'failed_checks',
            'critical_count', 'high_count', 'medium_count', 'low_count', 'status'
        ]
        
        # Format data
        formatted_data = []
        for scan in scans:
            formatted_data.append({
                'scan_id': scan.get('id', ''),
                'device_hostname': scan.get('device', {}).get('hostname', ''),
                'device_os_type': scan.get('device', {}).get('os_type', ''),
                'scan_date': scan.get('created_at', ''),
                'compliance_score': scan.get('compliance_score', 0),
                'total_checks': scan.get('total_checks', 0),
                'passed_checks': scan.get('passed_checks', 0),
                'failed_checks': scan.get('failed_checks', 0),
                'critical_count': scan.get('critical_count', 0),
                'high_count': scan.get('high_count', 0),
                'medium_count': scan.get('medium_count', 0),
                'low_count': scan.get('low_count', 0),
                'status': scan.get('status', ''),
            })
        
        return ExportService.generate_csv(formatted_data, headers)
